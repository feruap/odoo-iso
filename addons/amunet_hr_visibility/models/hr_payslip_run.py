import base64
import io
from odoo import models
from odoo.exceptions import UserError

# IDs de empleados que van en la lista de Nómina (más robusto que comparar nombres)
GRUPO1_IDS = {
    193,  # Stacy Abigail Palma Ramos
    195,  # Verónica Ortiz Moncada
    196,  # Alma Delia Bueno Garista
    197,  # Alondra Guadalupe Sánchez Martínez
    200,  # Verónica Natalia Perez Ruiz
    202,  # Patricia García Soto
    203,  # Diana Flores Vera
    209,  # Edgar Michel Salamanca Aguilar
}


class HrPayslipRun(models.Model):
    _inherit = 'hr.payslip.run'

    def action_generar_dispersion(self):
        try:
            import openpyxl
            from openpyxl.styles import Font
        except ImportError:
            raise UserError('Falta la librería openpyxl. Avisa a desarrollo.')

        grupo1, grupo2 = [], []

        for slip in self.slip_ids:
            emp = slip.employee_id
            net_line = slip.line_ids.filtered(lambda l: l.code == 'NET')
            neto = net_line[0].total if net_line else 0.0

            bank = self.env['res.partner.bank'].search(
                [('partner_id', '=', emp.work_contact_id.id)], limit=1
            )
            clabe = bank.acc_number if bank else ''
            banco = bank.bank_id.name if bank and bank.bank_id else ''

            row = (emp.name, banco, clabe, float(neto))
            if emp.id in GRUPO1_IDS:
                grupo1.append(row)
            else:
                grupo2.append(row)

        grupo1.sort(key=lambda x: x[0])
        grupo2.sort(key=lambda x: x[0])

        fecha = '%s al %s' % (
            self.date_start.strftime('%d/%m/%Y'),
            self.date_end.strftime('%d/%m/%Y'),
        )

        att_ids = []
        for label, filas in [('Nomina', grupo1), ('Honorarios', grupo2)]:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Lista 1 - Nomina'
            ws.column_dimensions['A'].width = 38
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 22
            ws.column_dimensions['D'].width = 12

            ws.append(['DISPERS\xc3\x93N — %s — %s' % (label.upper(), fecha),
                       None, None, None])
            ws['A1'].font = Font(bold=True, size=12)
            ws.append(['EMPLEADO', 'BANCO', 'CLABE', 'MONTO'])
            for col in ['A2', 'B2', 'C2', 'D2']:
                ws[col].font = Font(bold=True)

            for nombre, banco, clabe, monto in filas:
                ws.append([nombre, banco, clabe, monto])
                ws.cell(row=ws.max_row, column=3).number_format = '@'

            total = sum(r[3] for r in filas)
            ws.append([None, None, 'TOTAL', total])
            ws.cell(row=ws.max_row, column=3).font = Font(bold=True)
            ws.cell(row=ws.max_row, column=4).font = Font(bold=True)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

            nombre_archivo = 'Dispersion_%s_%s_%s.xlsx' % (
                label,
                self.date_start.strftime('%d%m%Y'),
                self.date_end.strftime('%d%m%Y'),
            )
            att = self.env['ir.attachment'].create({
                'name': nombre_archivo,
                'type': 'binary',
                'datas': base64.b64encode(buf.read()).decode(),
                'res_model': 'hr.payslip.run',
                'res_id': self.id,
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            })
            att_ids.append(att.id)

        return {
            'type': 'ir.actions.act_window',
            'name': 'Archivos de dispers\xf3n',
            'res_model': 'ir.attachment',
            'view_mode': 'list',
            'domain': [('id', 'in', att_ids)],
            'target': 'new',
        }
