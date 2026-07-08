import base64
import io
from odoo import models, fields
from odoo.exceptions import UserError


def _build_excel(label, filas, fecha, sin_clabe=None):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise UserError('Falta la librería openpyxl. Avisa a desarrollo.')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Lista 1 - Nomina'
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 12

    ws.append(['DISPERSIÓN — %s — %s' % (label.upper(), fecha), None, None, None])
    ws['A1'].font = Font(bold=True, size=12)
    ws.append(['EMPLEADO', 'BANCO', 'CLABE', 'MONTO'])
    for col in ['A2', 'B2', 'C2', 'D2']:
        ws[col].font = Font(bold=True)

    amarillo = PatternFill(fill_type='solid', fgColor='FFFF00')
    for nombre, banco, clabe, monto, nuevo_ingreso in filas:
        ws.append([nombre, banco, clabe, monto])
        ws.cell(row=ws.max_row, column=3).number_format = '@'
        if nuevo_ingreso:
            for col in range(1, 5):
                ws.cell(row=ws.max_row, column=col).font = Font(bold=True)
        if not clabe:
            ws.cell(row=ws.max_row, column=3).fill = amarillo

    total = sum(r[3] for r in filas)
    ws.append([None, None, 'TOTAL', total])
    ws.cell(row=ws.max_row, column=3).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=4).font = Font(bold=True)

    if sin_clabe:
        ws.append([])
        ws.append(['⚠ Sin CLABE (revisar antes de enviar al banco):', None, None, None])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color='FF0000')
        for nombre in sin_clabe:
            ws.append(['  • ' + nombre, None, None, None])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode()


class DispersionWizard(models.TransientModel):
    _name = 'amunet.dispersion.wizard'
    _description = 'Generar archivos de dispersión'

    nombre_nomina = fields.Char(string='Archivo Nómina')
    archivo_nomina = fields.Binary(string='Nómina', readonly=True)
    nombre_honorarios = fields.Char(string='Archivo Honorarios')
    archivo_honorarios = fields.Binary(string='Honorarios', readonly=True)
    resumen = fields.Char(string='Resumen', readonly=True)
    advertencia = fields.Char(string='Advertencia', readonly=True)

    def action_descargar_nomina(self):
        att = self.env['ir.attachment'].create({
            'name': self.nombre_nomina,
            'type': 'binary',
            'datas': self.archivo_nomina,
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%d?download=true' % att.id,
            'target': 'self',
        }

    def action_descargar_honorarios(self):
        att = self.env['ir.attachment'].create({
            'name': self.nombre_honorarios,
            'type': 'binary',
            'datas': self.archivo_honorarios,
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%d?download=true' % att.id,
            'target': 'self',
        }


class HrPayslipRun(models.Model):
    _inherit = 'hr.payslip.run'

    def action_generar_dispersion(self):
        self.ensure_one()
        grupo1, grupo2 = [], []
        sin_clabe_nomina, sin_clabe_honorarios = [], []

        # Empleados con ingreso en los últimos 30 días (para marcar en negritas)
        from datetime import date, timedelta
        corte_nuevo = date.today() - timedelta(days=30)

        for slip in self.slip_ids:
            emp = slip.employee_id
            net_line = slip.line_ids.filtered(lambda l: l.code == 'NET')
            neto = float(net_line[0].total) if net_line else 0.0

            bank = emp.bank_account_ids[:1] if emp.bank_account_ids else self.env['res.partner.bank']
            clabe = bank.acc_number if bank else ''
            banco = bank.bank_id.name if bank and bank.bank_id else ''

            # Nuevo ingreso: fecha de creación del empleado en los últimos 30 días
            nuevo = emp.create_date and emp.create_date.date() >= corte_nuevo

            row = (emp.name, banco, clabe, neto, nuevo)

            # Separar por estructura salarial del recibo (no por nombre hardcodeado)
            struct_name = (slip.struct_id.name or '').lower()
            if 'nomina' in struct_name or 'nómina' in struct_name:
                grupo1.append(row)
                if not clabe:
                    sin_clabe_nomina.append(emp.name)
            else:
                grupo2.append(row)
                if not clabe:
                    sin_clabe_honorarios.append(emp.name)

        grupo1.sort(key=lambda x: x[0])
        grupo2.sort(key=lambda x: x[0])

        fecha = '%s al %s' % (
            self.date_start.strftime('%d/%m/%Y'),
            self.date_end.strftime('%d/%m/%Y'),
        )
        sfecha = '%s_%s' % (
            self.date_start.strftime('%d%m%Y'),
            self.date_end.strftime('%d%m%Y'),
        )

        advertencia = ''
        todos_sin_clabe = sin_clabe_nomina + sin_clabe_honorarios
        if todos_sin_clabe:
            advertencia = '⚠ Sin CLABE: %s' % ', '.join(todos_sin_clabe)

        wizard = self.env['amunet.dispersion.wizard'].create({
            'nombre_nomina': 'Dispersion_Nomina_%s.xlsx' % sfecha,
            'archivo_nomina': _build_excel('Nómina', grupo1, fecha, sin_clabe_nomina),
            'nombre_honorarios': 'Dispersion_Honorarios_%s.xlsx' % sfecha,
            'archivo_honorarios': _build_excel('Honorarios', grupo2, fecha, sin_clabe_honorarios),
            'resumen': 'Nómina: %d empleados  |  Honorarios: %d empleados' % (
                len(grupo1), len(grupo2)),
            'advertencia': advertencia,
        })

        return {
            'type': 'ir.actions.act_window',
            'name': 'Archivos de dispersión',
            'res_model': 'amunet.dispersion.wizard',
            'res_id': wizard.id,
            'view_mode': 'form',
            'target': 'new',
        }
